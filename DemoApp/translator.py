import openpyxl
from googletrans import Translator

#class to translate file from one language to another
class language_translator:
    def __init__(self, doc, ln, filetype, filename, writepath):
        self.doc = doc
        self.ln = ln
        self.filetype = filetype
        self.filename = filename
        self.writepath = writepath
        self.translator = Translator(service_urls=[
        'translate.google.com',
        'translate.google.co.kr',
        ])


    def excel_translator(self):
        def column_finder():
            LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            column_list = []
            for letter in LETTERS:
                column_list.append(letter)
            for i in LETTERS:
                for j in LETTERS:
                    comb = i + j
                    column_list.append(comb)
            return column_list


        theFile = openpyxl.load_workbook(self.doc)
        allSheetNames = theFile.sheetnames

        print("All sheet names {} " .format(theFile.sheetnames))

        column_list = column_finder()
        for i in theFile.sheetnames:
            current_sheet = theFile[i]
            for row in range(1, current_sheet.max_row + 1):
                for column in range(0, current_sheet.max_column):
                    cell_value = current_sheet[f'{column_list[column]}{row}'].value
                    if type(cell_value) == str:
                        translated = self.translator.translate(cell_value, dest= self.ln)
                        current_sheet[f'{column_list[column]}{row}'].value = translated.text

        theFile.save(self.doc)
        print("Translation completed")
    
    def txt_translator(self):
        file1 = open(f'{self.doc}', 'r')
        file2 = open(f'{self.writepath}', 'w') 
        lines = file1.readlines()
        for line in lines:
            translated = self.translator.translate(line, dest= self.ln)
            file2.write(translated.text)
            file2.write("\n")
        file1.close()
        file2.close()
        return "Completed"

    def __repr__(self):
        return '{read_path: '+self.doc+', language: '+self.ln+', filetype: '+self.filetype+', filename: '+self.filename+', write_path: '+self.writepath+'}'

    def __str__(self):
        return 'language_translator(read_path='+self.doc+', language='+self.ln+', filetype='+self.filetype+', filename='+self.filename+', write_path='+self.writepath+'}'
